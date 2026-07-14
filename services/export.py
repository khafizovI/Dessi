import os
import re
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font

from config import EXPORTS_DIR
from database.repository import ReportData
from services.report_builder import fmt_money


def _exports_dir() -> str:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return str(EXPORTS_DIR)


def _pdf_safe(text: str) -> str:
    text = text.replace("\u2014", "-").replace("\u2013", "-")
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    text = text.replace("\u2026", "...")
    text = re.sub(r"[^\x00-\xFF]", "?", text)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _clean_label(label: str) -> str:
    for emoji in ("📅 ", "📆 ", "🗓 "):
        label = label.replace(emoji, "")
    return label


def export_excel(report: ReportData) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    bold = Font(bold=True)
    ws.append([_clean_label(report.period_label) + " hisobot"])
    ws["A1"].font = bold
    ws.append([])
    ws.append(["Jami tushum", report.revenue])
    ws.append(["Jami dastmoya", report.cost])
    ws.append(["Sotuv foydasi", report.profit])
    ws.append(["Xarajatlar", report.expenses])
    ws.append(["Sof foyda", report.net_profit])
    ws.append(["Sotilgan mahsulot", report.total_qty])
    ws.append(["Sotuvlar soni", report.sale_count])
    ws.append([])

    if report.admins:
        ws.append(["Sotuvchilar"])
        ws.append(["Ism", "Sotuvlar", "Dona", "Tushum"])
        for a in report.admins:
            ws.append([a["name"], a["count"], a["qty"], a["revenue"]])
        ws.append([])

    ws.append(["Mahsulot", "SKU", "Sotildi", "Tushum", "Dastmoya", "Foyda"])
    for p in report.products:
        ws.append([p["name"], p["sku"], p["qty"], p["revenue"], p["cost"], p["profit"]])

    filename = f"Hisobot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(_exports_dir(), filename)
    wb.save(filepath)
    return filepath


def export_pdf(report: ReportData) -> str:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)

    title = _pdf_safe(_clean_label(report.period_label))
    pdf.cell(0, 10, _pdf_safe(f"Dassi - {title} hisobot"), ln=True)
    pdf.ln(4)

    rows = [
        ("Jami tushum", fmt_money(report.revenue)),
        ("Jami dastmoya", fmt_money(report.cost)),
        ("Sotuv foydasi", fmt_money(report.profit)),
        ("Xarajatlar", fmt_money(report.expenses)),
        ("Sof foyda", fmt_money(report.net_profit)),
        ("Sotilgan mahsulot", str(report.total_qty)),
        ("Sotuvlar soni", str(report.sale_count)),
    ]
    for label, value in rows:
        pdf.cell(0, 8, _pdf_safe(f"{label}: {value}"), ln=True)

    if report.top_product:
        pdf.ln(4)
        pdf.cell(
            0, 8,
            _pdf_safe(
                f"Eng ko'p: {report.top_product['name']} "
                f"({report.top_product['qty']} ta)"
            ),
            ln=True,
        )
    if report.least_product and len(report.products) > 1:
        pdf.cell(
            0, 8,
            _pdf_safe(
                f"Eng kam: {report.least_product['name']} "
                f"({report.least_product['qty']} ta)"
            ),
            ln=True,
        )

    if report.admins:
        pdf.ln(4)
        pdf.set_font("Helvetica", style="B", size=11)
        pdf.cell(0, 8, "Sotuvchilar:", ln=True)
        pdf.set_font("Helvetica", size=10)
        for a in report.admins:
            pdf.cell(
                0, 7,
                _pdf_safe(
                    f"{a['name']}: {a['count']} sotuv, "
                    f"{fmt_money(a['revenue'])} so'm"
                ),
                ln=True,
            )

    if report.products:
        pdf.ln(4)
        pdf.set_font("Helvetica", style="B", size=11)
        pdf.cell(0, 8, "Mahsulotlar:", ln=True)
        pdf.set_font("Helvetica", size=9)
        for p in report.products:
            pdf.cell(
                0, 6,
                _pdf_safe(
                    f"{p['name']} | {p['qty']} ta | "
                    f"{fmt_money(p['revenue'])} | "
                    f"foyda {fmt_money(p['profit'])}"
                ),
                ln=True,
            )

    filename = f"Hisobot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(_exports_dir(), filename)
    pdf.output(filepath)
    return filepath
